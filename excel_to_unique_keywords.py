import os
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

extension = sys.argv[2]

wb = load_workbook(filename=sys.argv[1], read_only=True)
ws = wb.active
unique_keywords = []
i=0
rows = list(ws.rows)
while (i < ws.max_row) and (rows[i][0].value != None):
    keyword = str(rows[i][1].value.encode('utf-8'))
    if keyword not in unique_keywords:
        unique_keywords.append(keyword)
    i = i + 1

newWB = Workbook()
newWS = newWB.active
n = 1
for elt in unique_keywords:
    _ = newWS.cell(column=1, row=n, value=1)
    _ = newWS.cell(column=2, row=n, value=elt)
    n = n + 1

work_dir = os.getcwd() + "/corpus/"
base = os.path.basename(sys.argv[1])
filename = os.path.splitext(base)[0]
newWB.save(work_dir + filename  + extension)
