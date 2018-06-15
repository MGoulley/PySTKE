#encoding: utf-8

import os
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import nltk
from openpyxl.styles import Color, PatternFill, Font, Border
from PyPDF2 import PdfFileReader, PdfFileWriter

def isAproxIn(value, table):
    for elt in table:
        if (value in elt) or (elt in value):
            return True
    return False

def getAprox(value, table):
    for elt in table:
        if (value in elt) or (elt in value):
            return elt


def get_kw_n(n, docpath):
    i=0
    docwb = load_workbook(filename=docpath, read_only=True)
    # docws = docwb['Sheet1']
    docws = docwb.active
    element_of_slide = []
    rows = list(docws.rows)
    while (i < docws.max_row) and (rows[i][0].value != None):
        if(int(rows[i][0].value) == n):
            element_of_slide.append(rows[i][1].value)
        i = i + 1
    return element_of_slide

def compute_results(tab1, tab2):
    y_true = np.array(tab1)
    y_pred = np.array(tab2)
    multi = MultiLabelBinarizer()
    A_new = multi.fit_transform(y_true)
    B_new = multi.transform(y_pred)
    precision_score(A_new,B_new,average='samples')
    recall_score(A_new, B_new, average='samples')
    return precision_recall_fscore_support(A_new, B_new, average='samples')

refdocwb = load_workbook(filename=sys.argv[1], read_only=True)
refdocdict = {}
#  refdocws = refdocwb['Sheet1']
refdocws = refdocwb.active
element_of_slide = []
key = 1
prev_key = 0
i=0
rows = list(refdocws.rows)
while (i < refdocws.max_row) and (rows[i][0].value != None):
    key = int(rows[i][0].value or 0)
    if key > prev_key:
        refdocdict[prev_key] = element_of_slide
        element_of_slide = []
        prev_key = key
    element_of_slide.append(rows[i][1].value.rstrip().lower())
    i = i + 1
refdocdict[prev_key] = element_of_slide



compdocwb = load_workbook(filename=sys.argv[2], read_only=True)
compdocdict = {}
# compdocws = compdocwb['Sheet1']
compdocws = compdocwb.active
element_of_slide = []
key = 1
prev_key = 0
i=0
rows = list(compdocws.rows)
while (i < compdocws.max_row) and (rows[i][1].value != None):
    key = int(rows[i][0].value or 0)
    if key > prev_key:
        compdocdict[prev_key] = element_of_slide
        element_of_slide = []
        prev_key = key
    element_of_slide.append(rows[i][1].value.rstrip().lower())
    i = i + 1
compdocdict[prev_key] = element_of_slide


wb = Workbook()
ws = wb.active
actual_row = 1

max_key = 0
for key in refdocdict.keys():
    if key > max_key:
        max_key = key

for key in compdocdict.keys():
    if key > max_key:
        max_key = key

size_ref = sum(map(len, refdocdict.values()))
size_comp = sum(map(len, compdocdict.values()))
match = 0
err = 0
approx = 0


act_key = 0
while act_key <= max_key: # Toutes les clés jusqu'a la dernière existante
    if (act_key in compdocdict.keys()) and (act_key in refdocdict.keys()): # si la clé actuelle est dans les deux dict
        save_approx = []
        for refvalue in refdocdict[act_key]: # parcours des valeurs dans le dictionnaire de reference
            if refvalue in compdocdict[act_key]: # Si une valeure du dict ref est dans le dict de comparaison exactement
                _ = ws.cell(column=1, row=actual_row, value=act_key)
                _ = ws.cell(column=2, row=actual_row, value=refvalue)
                _ = ws.cell(column=3, row=actual_row, value=refvalue)
                _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="33CC33", end_color="33CC33", fill_type = "solid")
                actual_row = actual_row + 1
                compdocdict[act_key].remove(refvalue)
                match = match +1
            elif isAproxIn(value=refvalue, table=compdocdict[act_key]): # Si la valeur du dict ref ressemble approximativement a une expression
                val = getAprox(value=refvalue, table=compdocdict[act_key])
                _ = ws.cell(column=1, row=actual_row, value=act_key)
                _ = ws.cell(column=2, row=actual_row, value=refvalue)
                _ = ws.cell(column=3, row=actual_row, value=val)
                _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type = "solid")
                actual_row = actual_row + 1
                save_approx.append(val)
                approx = approx + 1
            else:
                _ = ws.cell(column=1, row=actual_row, value=act_key)
                _ = ws.cell(column=2, row=actual_row, value=refvalue)
                _ = ws.cell(column=3, row=actual_row)
                _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type = "solid")
                actual_row = actual_row + 1
                err = err + 1
        for compvalue in compdocdict[act_key]: # parcours des valeurs dans le dictionnaire de comparaison (forcement introuvable dans le dictionnaire de ref)
            if compvalue not in save_approx:
                _ = ws.cell(column=1, row=actual_row, value=act_key)
                _ = ws.cell(column=2, row=actual_row)
                _ = ws.cell(column=3, row=actual_row, value=compvalue)
                _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type = "solid")
                actual_row = actual_row + 1
                err = err + 1
    elif (act_key in compdocdict.keys()) and (act_key not in refdocdict.keys()): # si la clé actuelle est dans le dictionnaire de comparaison
        for compvalue in compdocdict[act_key]:
            _ = ws.cell(column=1, row=actual_row, value=act_key)
            _ = ws.cell(column=2, row=actual_row)
            _ = ws.cell(column=3, row=actual_row, value=compvalue)
            _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type = "solid")
            actual_row = actual_row + 1
            err = err + 1
    elif (act_key not in compdocdict.keys()) and (act_key in refdocdict.keys()): # si la clé actuelle est dans le dictionnaire de référence
        for refvalue in refdocdict[act_key]:
            _ = ws.cell(column=1, row=actual_row, value=act_key)
            _ = ws.cell(column=2, row=actual_row, value=refvalue)
            _ = ws.cell(column=3, row=actual_row)
            _ = ws.cell(column=4, row=actual_row).fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type = "solid")
            actual_row = actual_row + 1
            err = err + 1
    act_key = act_key +1



base = os.path.basename(sys.argv[2])
filename = os.path.splitext(base)[0]
work_dir = os.getcwd() + "/corpus/"
wb.save(work_dir + "comparaison_"+ filename  +".xlsx")
print("Il y a " + str(match) + " matchs, "+ str(err) + " erreurs, " + str(approx) + " approximations.")
print("Nombre d'éléments annotés automatiquement : " + str(size_comp))
print("Nombre d'éléments annotés manuellement : " + str(size_ref))
precision = (match+approx)/float(size_comp)
rappel = (match+approx)/float(size_ref)
fscore = 2*(precision*rappel)/(precision+rappel)
print("Precision: " + str(precision) + " Rappel: " + str(rappel) + " F-Score: " + str(fscore))


# UTILISATION
# python3 exel_comp.py /home/matthias/rapportstage/annotations/annotations_cours_fabrice_matthias.xlsx /home/matthias/pyKE/corpus/TICA.xlsx
# python3 exel_comp.py /home/matthias/rapportstage/annotations/annotations_cours_fabrice_matthias.xlsx /home/matthias/pyKE/corpus/TICA.xlsx /home/matthias/pyKE/corpus/TICA.pdf
