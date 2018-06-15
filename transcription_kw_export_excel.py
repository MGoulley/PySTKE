import os
import sys
from openpyxl import Workbook

# Ecriture dans fichier excel
fichier = sys.argv[1]
dir = sys.argv[2]
work_dir = os.getcwd() + "/corpus/"
base = os.path.basename(dir)
filename = os.path.splitext(base)[0]


wb = Workbook()
ws = wb.active
if os.path.isfile(fichier):
    file = open(fichier,"r")
    kw_lst = list(file)
    nb = 0
    for elt in kw_lst:
        _ = ws.cell(column=1, row=nb+1, value=1)
        _ = ws.cell(column=2, row=nb+1, value=elt.rstrip().encode('utf-8'))
        nb = nb + 1
    wb.save(work_dir + filename + ".xlsx")
