#!/usr/bin/python3
#encoding: utf-8

"""
Programme permetant d'extraire des mots clés à partir d'un
diaporama ou d'une transcription de la parole.
"""

import os
import os.path
import sys
import argparse
import re
from openpyxl import load_workbook

# Fonctions
def count_words_in_text_file(file_path):
    """
        Compte le nombre de mots dans un document texte.
    """
    if os.path.isfile(file_path):
        file_path = open(file_path, "r", encoding="ISO-8859-1")
        raw_text = list(file_path)
        nb_words_in_document = 0
        for element in raw_text:
            element = element.rstrip()
            element = re.sub(r'.*> ', '', element)
            sentence = element.split()
            nb_words_in_document = nb_words_in_document + len(sentence)
        print("Il y a dans ce document " + str(nb_words_in_document) + " mots.")
        file_path.close()
    return nb_words_in_document

def count_words_in_table_file(file_path):
    """
        Compte le nombre de mots dans les expression annotées manuellement.
    """
    workbook = load_workbook(filename=file_path, read_only=True)
    worksheet = workbook.active
    nb_words_in_document = 0
    i = 0
    rows = list(worksheet.rows)
    while (i < worksheet.max_row) and (rows[i][0].value != None):
        element = rows[i][1].value.rstrip().lower()
        sentence = element.split()
        nb_words_in_document = nb_words_in_document + len(sentence)
        i = i + 1
    print("Vous avez annoté manuellement " + str(nb_words_in_document) + " mots.")
    return nb_words_in_document

def str2bool(chaine):
    """
        Transforme une chaine de caractère en booléen.
    """
    if chaine.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif chaine.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')

def keywords_extraction(tol, wrkdir, filnam):
    """
        Extrait les mots clés d'un document avec l'outil passé en paramètre.
    """
    if tol == 'pke':
        os.system('python3 txt_to_kw_pke.py ' + wrkdir + filnam + ".txt")
    elif tol == 'pyrata':
        os.system('python3 txt_to_kw_pyrata.py ' + wrkdir + filnam + ".txt")
    elif tol == 'mixt':
        os.system('python3 txt_to_kw_mixt.py ' + wrkdir + filnam + ".txt")
    elif tol == 'wikifier':
        os.system('python3 txt_to_kw_wikifier.py ' + wrkdir + filnam + ".txt")
    elif os.path.isfile('txt_to_kw_' + tol + '.py'):
        os.system('python3 txt_to_kw_' + tol + '.py ' + wrkdir + filnam + ".txt")
    else:
        sys.exit("L'outil " + tol + " n'existe pas.")

def comparaison_par_diapositive(file_ref, file_comp):
    """
        Compare deux fichier de mots clés par diapositive.
    """
    if os.path.isfile(file_ref) and os.path.isfile(file_comp):
        base_ref = os.path.basename(file_ref)
        extension_ref = os.path.splitext(base_ref)[1]
        if extension_ref == '.xlsx':
            os.system('python excel_comp.py '+ file_ref +' ' + file_comp)
        else:
            sys.exit("L'extension du fichier " + file_ref + " n'est pas conforme.")
    else:
        sys.exit("Le fichier " + file_ref + " n'existe pas.")

def comparaison_par_mot_cle_unique(file_ref, file_comp):
    """
        Compare deux fichier de mots clés par expression unique.
    """
    work_dir = os.getcwd() + "/corpus/"
    if os.path.isfile(file_ref):
        base_ref = os.path.basename(file_ref)
        filename_ref = os.path.splitext(base_ref)[0]
        extension_ref = os.path.splitext(base_ref)[1]
        if extension_ref == '.xlsx':
            os.system('python excel_to_unique_keywords.py '+ file_ref + ' _unique_ref.xlsx')
        else:
            sys.exit("L'extension du fichier " + file_ref + " n'est pas conforme.")
    else:
        sys.exit("Le fichier " + file_ref + " n'existe pas.")
    if os.path.isfile(file_comp):
        base_comp = os.path.basename(file_comp)
        filename_comp = os.path.splitext(base_comp)[0]
        extension_comp = os.path.splitext(base_comp)[1]
        if extension_comp == '.xlsx':
            os.system('python excel_to_unique_keywords.py '+ file_comp + ' _unique_comp.xlsx')
        else:
            sys.exit("L'extension du fichier " + file_comp + " n'est pas conforme.")
    else:
        sys.exit("Le fichier " + file_comp + " n'existe pas.")
    comparaison_par_diapositive(work_dir + filename_ref +'_unique_ref.xlsx', work_dir +
                                filename_comp + '_unique_comp.xlsx')

# Analyse des arguments de lancement
PARSER = argparse.ArgumentParser(description='Extracteur automatique de mots clés')
PARSER.add_argument('-d', '--doc',
                    help='Chemin vers le document à annoter automatiquement')
PARSER.add_argument('-t', '--tool', nargs='?', default='pyrata',
                    help='Outil à utiliser pour extraire les mots clés:'
                    'pke, pyrata, mixt, wikifier (default pyrata)')
PARSER.add_argument('-r', '--ref', nargs='?', help='Chemin vers le document d\'annotation'
                    ' manuelle à comparer avec l\'extraction automatique')
PARSER.add_argument("--unique", type=str2bool, nargs='?', const=True, default=False,
                    help="Comparaison par mots clés uniques dans le document. "
                    "(Defaut : Par diapositive)")
PARSER.add_argument("--count", type=str2bool, nargs='?', const=True, default=False,
                    help="Ajoute des détails sur les expressions annotées manuellement")

ARGS = PARSER.parse_args()
DIR = ARGS.doc
FILE_REF = ARGS.ref
TOOL = ARGS.tool
UNIQUE = ARGS.unique
COUNT = ARGS.count

# Constantes
EXIDE_SUPPORT = ['.pptx', '.odp', '.tex'] # .ppt ?
TRANSCRIPTION_SUPPORT = ['.stm', '.txt']

# Début du programme
if DIR != None:
    # On stockera tous les fichiers produits dans un repertoire corpus
    WORK_DIR = os.getcwd() + "/corpus/"
    BASE = os.path.basename(DIR)
    FILENAME = os.path.splitext(BASE)[0]
    EXTENSION = os.path.splitext(BASE)[1]
    if os.path.isfile(DIR):
        if EXTENSION in EXIDE_SUPPORT: # Cas d'un fichier source
            # Step 1 : Extraction du texte
            os.system('python extract_src.py ' + DIR + " " + WORK_DIR)
            # Step 2 : Extraction des mots clés
            keywords_extraction(TOOL, WORK_DIR, FILENAME)
            # Step 3 : Alignement des mots clés par diapositive
            os.system('python src_kw_export_excel.py ' + WORK_DIR + FILENAME + "_kw.txt " + DIR)
        elif EXTENSION == '.pdf': # Cas d'un fichier PDF
            # Step 1 : Extraction du texte
            os.system('python extract_pdf.py ' + DIR + " " + WORK_DIR)
            # Step 2 : Extraction des mots clés
            keywords_extraction(TOOL, WORK_DIR, FILENAME)
            # Step 3 : Alignement des mots clés par diapositive
            os.system('python3 pdf_kw_export_excel.py ' + WORK_DIR + FILENAME + "_kw.txt " + DIR)
        elif EXTENSION in TRANSCRIPTION_SUPPORT: # Cas d'une Transcription
            # Step 1 : Extraction du texte
            os.system('python3 extract_transcription.py ' + DIR + " " + WORK_DIR)
            # Step 2 : Extraction des mots clés
            keywords_extraction(TOOL, WORK_DIR, FILENAME)
            # Step 3 : Alignement des mots clés
            os.system('python3 transcription_kw_export_excel.py ' + WORK_DIR + FILENAME
                      + "_kw.txt " + DIR)
        else:
            sys.exit("L'extension du fichier " + DIR + " n'est pas conforme.")
    else:
        sys.exit("Le fichier " + DIR + " n'existe pas.")
    # Step 4 : Comparaison des résultats avec la ref dans un tableur
    if FILE_REF != None:
        if UNIQUE:
            comparaison_par_mot_cle_unique(FILE_REF, WORK_DIR + FILENAME + '.xlsx')
        else:
            comparaison_par_diapositive(FILE_REF, WORK_DIR + FILENAME + '.xlsx')
    # Step 5 : Résultats complémentaires annotation manuelle
    if COUNT:
        WORDS_COUNT = count_words_in_text_file(WORK_DIR + FILENAME + ".txt")
        if FILE_REF != None:
            WORDS_ANNOTATE = count_words_in_table_file(FILE_REF)
            RESULT = (WORDS_ANNOTATE * 100)/WORDS_COUNT
            print("Cela représente " + "{0:.3f}".format(RESULT) + "%% de mots annotés.")
else:
    print("Vous n'avez pas spécifier de fichier à analyser.")
