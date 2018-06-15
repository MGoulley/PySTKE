import os
import sys
from openpyxl import Workbook
from PyPDF2 import PdfFileReader, PdfFileWriter


# Ecriture dans fichier excel
fichier = sys.argv[1]
dir = sys.argv[2]
work_dir = os.getcwd() + "/corpus/"
base = os.path.basename(dir)
filename = os.path.splitext(base)[0]


inputpdf = PdfFileReader(open(dir, "rb"))
wb = Workbook()
ws = wb.active
if os.path.isfile(fichier):
    file = open(fichier,"r")
    kw_lst = list(file)
    nb = 0
    for i in range(inputpdf.numPages):
        txt = inputpdf.getPage(i).extractText().rstrip().encode('utf-8')
        for word in kw_lst:
            if word.rstrip().encode('utf-8') in txt:
                _ = ws.cell(column=1, row=nb+1, value=i+1)
                _ = ws.cell(column=2, row=nb+1, value=word.rstrip().encode('utf-8'))
                nb = nb + 1
    wb.save(work_dir + filename + ".xlsx")
