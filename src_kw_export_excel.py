from exide.parse import parse
import os
import sys
from openpyxl import Workbook


# Ecriture dans fichier excel
fichier = sys.argv[1]
dir = sys.argv[2]
exide_presentation = parse(dir)
work_dir = os.getcwd() + "/corpus/"

if os.path.isfile(fichier):
    file = open(fichier,"r")
    list_txt = list(file)
    i = 1
    data = ""
    wb = Workbook()
    ws = wb.active
    index = 0
    while exide_presentation.get_slide_by_id(i) != None:
        data = exide_presentation.get_slide_by_id(i).text
        for elt in list_txt:
            if elt.decode('utf-8').rstrip() in data:
                _ = ws.cell(column=1, row=index+1, value=i)
                _ = ws.cell(column=2, row=index+1, value=elt.decode('utf-8').rstrip())
                index = index + 1
        i = i+1
    file.close()
    base = os.path.basename(dir)
    filename = os.path.splitext(base)[0]
    wb.save(work_dir + filename + ".xlsx")
